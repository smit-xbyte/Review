import os
from openpyxl import load_workbook
from openpyxl import Workbook

def save_to_excel(data):
    file_path = "employee_reviews.xlsx"

    # Check if the file exists
    if os.path.exists(file_path):
        # If the file exists, load the workbook and the sheet
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # If the file does not exist, create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Reviews"
        
        # Define the header and append it to the first row
        headers = [
            'Developer Name', 'Emp ID', 'Review Month', 'Team Leader', 'Project Manager',
            'Delivery Timelines', 'Utilization', 'Technology Understanding', 'Attitude Responsibility',
            'Code Quality', 'Communication Skills', 'Glacier Process Alignment', 'Bugs Iterations',
            'Attendance', 'Outstanding Performance', 'Overall Rating', 'Remarks'
        ]
        ws.append(headers)

    # Append data from the review
    for record in data:
        row = [
            record.developer_name, record.emp_id, record.review_month, record.team_leader,
            record.project_manager, record.delivery_timelines, record.utilization,
            record.technology, record.attitude, record.code_quality,
            record.communication, record.glacier_process, record.bugs,
            record.attendance, record.outstanding, record.rating, record.remarks
        ]
        ws.append(row)

    # Save the workbook (this will overwrite the existing file)
    wb.save(file_path)

    return file_path
